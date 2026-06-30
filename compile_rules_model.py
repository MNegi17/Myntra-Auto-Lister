import os
import json
import pandas as pd
import openpyxl

learning_dir = r"C:\Users\Manann\Downloads"
ballerinas_file = r"C:\Users\Manann\Downloads\Myntra_Listing-Sheets(11-03-26)\Myntra-Sku-Template-2026-03-11_Ballerinas.xlsx"
top_file = os.path.join(learning_dir, "Myntra-Sku-Template-2026-05-05-Tops.xlsx")
sizechart_file = os.path.join(learning_dir, "Size Chart (4).xlsx")

rules_db = {
    "apparel_defaults": {},
    "footwear_defaults": {},
    "apparel_sizecharts": {},
    "footwear_sizecharts": {}
}

def normalize_size_label(size):
    s = str(size).strip().upper()
    s = s.replace('/', '-')  # 2/3Y -> 2-3Y
    s = s.replace(' ', '')
    
    # Custom numeric size mapping to standard kids age groups
    numeric_mappings = {
        "2": "2-3Y",
        "2Y": "2-3Y",
        "3": "3-4Y",
        "3Y": "3-4Y",
        "4": "3-4Y",
        "4Y": "3-4Y",
        "5": "4-5Y",
        "5Y": "4-5Y",
        "6": "5-6Y",
        "6Y": "5-6Y",
        "7": "7-8Y",
        "7Y": "7-8Y",
        "8": "7-8Y",
        "8Y": "7-8Y",
        "9": "9-10Y",
        "9Y": "9-10Y",
        "10": "9-10Y",
        "10Y": "9-10Y",
        "11": "11-12Y",
        "11Y": "11-12Y",
        "12": "11-12Y",
        "12Y": "11-12Y",
        "13": "13-14Y",
        "13Y": "13-14Y",
        "14": "13-14Y",
        "14Y": "13-14Y"
    }
    
    if s in numeric_mappings:
        return numeric_mappings[s]
        
    return s

# 1. Parse Stacked Sizing Chart (Size Charts.xlsx)
print("=== Parsing Stacked Size Charts ===")
if os.path.exists(sizechart_file):
    df_chart = pd.read_excel(sizechart_file, header=None)
    current_category = None
    sizes = []
    
    for idx, row in df_chart.iterrows():
        val = row[0]
        # Detect table headers (single column text in row)
        if pd.notna(val) and isinstance(val, str) and not val.startswith("POM IN") and "INSEAM" not in val.upper() and row[1:].isna().all():
            current_category = val.strip().upper()
            rules_db["apparel_sizecharts"][current_category] = {}
            sizes = []
            continue
            
        if current_category is None:
            continue
            
        # Detect sizes row
        if pd.notna(val) and isinstance(val, str) and val.startswith("POM IN"):
            sizes = [normalize_size_label(x) for x in row[2:] if pd.notna(x)]
            continue
            
        # Parse measurement cell
        if pd.notna(val) and len(sizes) > 0:
            measurement_name = str(val).strip()
            for col_idx, size in enumerate(sizes):
                if col_idx + 2 < len(row):
                    val_cell = row[col_idx + 2]
                    if pd.notna(val_cell):
                        if size not in rules_db["apparel_sizecharts"][current_category]:
                            rules_db["apparel_sizecharts"][current_category][size] = {}
                        # Keep it standard
                        rules_db["apparel_sizecharts"][current_category][size][measurement_name] = val_cell
                        
    print(f"Parsed {len(rules_db['apparel_sizecharts'])} Apparel categories from Size Charts.")
else:
    print("WARNING: Size Charts file not found!")

# 2. Parse Footwear Sizes directly from Footwear Reference Sheet
print("\n=== Parsing Footwear Reference Sizes ===")
if os.path.exists(ballerinas_file):
    wb = openpyxl.load_workbook(ballerinas_file, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s not in ["__INSTRUCTIONS", "masterdata"]]
    if sheets:
        sheet_name = sheets[0]
        df_pd = pd.read_excel(ballerinas_file, sheet_name=sheet_name, header=None)
        if df_pd.shape[0] >= 3:
            tech_keys = [str(k).strip() for k in df_pd.iloc[2]]
            data_df = df_pd.iloc[3:].copy()
            data_df.columns = [k if pd.notna(k) and str(k).strip() != 'nan' else f"Col_{i}" for i, k in enumerate(tech_keys)]
            
            size_col = 'Brand Size' if 'Brand Size' in data_df.columns else ('Standard Size' if 'Standard Size' in data_df.columns else None)
            if size_col:
                measurement_cols = [c for c in data_df.columns if 'inches' in c.lower() or 'cm' in c.lower() or 'length' in c.lower() or 'size' in c.lower() or 'age-group' in c.lower()]
                
                # We save Footwear Sizecharts under BALLERINAS / FOOTWEAR
                rules_db["footwear_sizecharts"]["FOOTWEAR"] = {}
                for idx, row in data_df.iterrows():
                    sz = str(row[size_col]).strip().upper()
                    if pd.notna(sz) and sz != 'nan' and sz not in rules_db["footwear_sizecharts"]["FOOTWEAR"]:
                        rules_db["footwear_sizecharts"]["FOOTWEAR"][sz] = {}
                        for col in measurement_cols:
                            if col != size_col and pd.notna(row[col]) and str(row[col]).strip() != 'nan':
                                rules_db["footwear_sizecharts"]["FOOTWEAR"][sz][col] = row[col]
                                
                print(f"Extracted {len(rules_db['footwear_sizecharts']['FOOTWEAR'])} footwear sizes from reference.")
    wb.close()

# 3. Parse and Override Apparel Static Defaults (紫 Kids)
print("\n=== Parsing Apparel Static Defaults ===")
if os.path.exists(top_file):
    wb = openpyxl.load_workbook(top_file, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s not in ["__INSTRUCTIONS", "masterdata"]]
    if sheets:
        sheet_name = sheets[0]
        df_pd = pd.read_excel(top_file, sheet_name=sheet_name, header=None)
        if df_pd.shape[0] >= 3:
            tech_keys = [str(k).strip() for k in df_pd.iloc[2]]
            data_df = df_pd.iloc[3:].copy()
            data_df.columns = [k if pd.notna(k) and str(k).strip() != 'nan' else f"Col_{i}" for i, k in enumerate(tech_keys)]
            
            candidate_cols = [
                'Manufacturer Name and Address with Pincode', 'Packer Name and Address with Pincode',
                'Importer Name and Address with Pincode', 'Country Of Origin', 'is Standard Size present on Label',
                'FashionType', 'Usage', 'Fabric', 'Fit', 'Pattern', 'Occasion', 'Knit or Woven', 'Closure',
                'Wash Care', 'Multipack Set', 'Fabric Type', 'Number of Items', 'Net Quantity Unit', 
                'Package Contains', 'Net Quantity', 'materialCareDescription', 'Theme'
            ]
            for col in candidate_cols:
                if col in data_df.columns:
                    non_empty = data_df[col].dropna()
                    if not non_empty.empty:
                        rules_db["apparel_defaults"][col] = non_empty.mode()[0]
                        
    # Hardcode company rule for Apparel brand
    rules_db["apparel_defaults"]["brand"] = "Purple United Kids"
    print("Apparel brand hardcoded to: Purple United Kids")
    wb.close()

# 4. Parse and Override Footwear Static Defaults (toothless)
print("\n=== Parsing Footwear Static Defaults ===")
if os.path.exists(ballerinas_file):
    wb = openpyxl.load_workbook(ballerinas_file, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s not in ["__INSTRUCTIONS", "masterdata"]]
    if sheets:
        sheet_name = sheets[0]
        df_pd = pd.read_excel(ballerinas_file, sheet_name=sheet_name, header=None)
        if df_pd.shape[0] >= 3:
            tech_keys = [str(k).strip() for k in df_pd.iloc[2]]
            data_df = df_pd.iloc[3:].copy()
            data_df.columns = [k if pd.notna(k) and str(k).strip() != 'nan' else f"Col_{i}" for i, k in enumerate(tech_keys)]
            
            candidate_cols = [
                'Manufacturer Name and Address with Pincode', 'Packer Name and Address with Pincode',
                'Importer Name and Address with Pincode', 'Country Of Origin', 'is Standard Size present on Label',
                'FashionType', 'Usage', 'Material', 'Toe Shape', 'Pattern', 'Occasion', 'Insole', 'Type',
                'Wash Care', 'Multipack Set', 'Number of Items', 'Net Quantity Unit', 'Package Contains', 
                'Net Quantity', 'materialCareDescription', 'Theme', 'Fastening', 'Cleats', 'Soles'
            ]
            for col in candidate_cols:
                if col in data_df.columns:
                    non_empty = data_df[col].dropna()
                    if not non_empty.empty:
                        rules_db["footwear_defaults"][col] = non_empty.mode()[0]
                        
    # Hardcode company rule for Footwear brand
    rules_db["footwear_defaults"]["brand"] = "toothless"
    print("Footwear brand hardcoded to: toothless")
    wb.close()

# 5. Hardcode new Manufacturer and Packer Address
new_address = "Purple United Sales Limited, Khasra No. 55/14 & 55/15, Mundka, Near Rani Khera Road, New Delhi, West Delhi, Delhi 110041."
rules_db["apparel_defaults"]["Manufacturer Name and Address with Pincode"] = new_address
rules_db["apparel_defaults"]["Packer Name and Address with Pincode"] = new_address
rules_db["footwear_defaults"]["Manufacturer Name and Address with Pincode"] = new_address
rules_db["footwear_defaults"]["Packer Name and Address with Pincode"] = new_address

# 6. Inject custom JEANS size charts for boys and girls
rules_db["apparel_sizecharts"]["JEANS"] = {
    "2-3Y": {
        "Inseam Length ( Inches )": 17.3,
        "To Fit Waist ( Inches )": 12.6
    },
    "3-4Y": {
        "Inseam Length ( Inches )": 18.9,
        "To Fit Waist ( Inches )": 13.2
    },
    "4-5Y": {
        "Inseam Length ( Inches )": 19.7,
        "To Fit Waist ( Inches )": 14.4
    },
    "5-6Y": {
        "Inseam Length ( Inches )": 20.5,
        "To Fit Waist ( Inches )": 15.8
    },
    "7-8Y": {
        "Inseam Length ( Inches )": 22.1,
        "To Fit Waist ( Inches )": 16.9
    },
    "9-10Y": {
        "Inseam Length ( Inches )": 23.6,
        "To Fit Waist ( Inches )": 18.9
    },
    "11-12Y": {
        "Inseam Length ( Inches )": 25.2,
        "To Fit Waist ( Inches )": 21.6
    },
    "13-14Y": {
        "Inseam Length ( Inches )": 26.8,
        "To Fit Waist ( Inches )": 24.4
    }
}

# Write workspace JSON rules model database
output_paths = [
    "myntra_rules_db.json",
    r"c:\Users\Manann\Desktop\Myntra Auto Lister V2\myntra_rules_db.json",
    r"c:\Users\Manann\Desktop\Myntra Auto Lister\myntra_rules_db.json"
]

for path in output_paths:
    try:
        dir_name = os.path.dirname(path)
        if dir_name and not os.path.exists(dir_name):
            os.makedirs(dir_name, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(rules_db, f, indent=4, default=str)
        print(f"[AI SUCCESS] Saved rules database to: {path}")
    except Exception as e:
        print(f"WARNING: Could not save to {path}: {str(e)}")

