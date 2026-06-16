import os
import sys
import json
import time
import queue
import threading
import traceback
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from flask import Flask, jsonify, request, Response, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
import vision_classifier
try:
    import tkinter as tk
    from tkinter import filedialog
except ImportError:
    tk = None
    filedialog = None

# Load environment variables from .env file if it exists
if os.path.exists(".env"):
    try:
        with open(".env", "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip()
    except Exception as e:
        print(f"Error loading .env file: {e}")

IS_CLOUD = os.environ.get("IS_CLOUD", "false").lower() == "true" or tk is None or filedialog is None

# Load AI learned offline rules database
rules_db = {}
rules_file = "myntra_rules_db.json"
if os.path.exists(rules_file):
    try:
        with open(rules_file, "r", encoding="utf-8") as f:
            rules_db = json.load(f)
        print(f"[AI SUCCESS] Loaded offline rules database from: {rules_file}")
    except Exception as e:
        print(f"ERROR: Failed to load rules database: {str(e)}")
else:
    print(f"WARNING: Sizing and rules database '{rules_file}' not found. Please compile it first.")

def normalize_size_label(size):
    s = str(size).strip().upper()
    s = s.replace('/', '-')  # 2/3Y -> 2-3Y
    s = s.replace(' ', '')
    
    # Custom numeric size mapping to standard kids age groups
    numeric_mappings = {
        "2": "2-3Y",
        "2Y": "2-3Y",
        "3": "3-4Y",
        "3Y": "3-4Y",
        "4": "3-4Y",
        "4Y": "3-4Y",
        "5": "4-5Y",
        "5Y": "4-5Y",
        "6": "5-6Y",
        "6Y": "5-6Y",
        "7": "7-8Y",
        "7Y": "7-8Y",
        "8": "7-8Y",
        "8Y": "7-8Y",
        "9": "9-10Y",
        "9Y": "9-10Y",
        "10": "9-10Y",
        "10Y": "9-10Y",
        "11": "11-12Y",
        "11Y": "11-12Y",
        "12": "11-12Y",
        "12Y": "11-12Y",
        "13": "13-14Y",
        "13Y": "13-14Y",
        "14": "13-14Y",
        "14Y": "13-14Y"
    }
    
    if s in numeric_mappings:
        return numeric_mappings[s]
        
    return s

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

# Global variables for task tracking and logs
current_task_status = {
    "status": "idle",
    "progress": 0.0,
    "current_step": "",
    "error": None
}

class LogBuffer:
    def __init__(self):
        self.listeners = []

    def add_listener(self):
        q = queue.Queue()
        self.listeners.append(q)
        return q

    def remove_listener(self, q):
        if q in self.listeners:
            self.listeners.remove(q)

    def log(self, level, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = {
            "time": timestamp,
            "level": level,
            "message": message
        }
        for q in self.listeners:
            q.put(log_entry)

log_buffer = LogBuffer()

# Reused Categories and Defaults from Tkinter app
CATEGORY_DEFAULTS = {
    "shorts": {
        "Package Contains": "1 Short",
        "Type": "Regular Shorts",
        "Length": "Above Knee"
    },
    "trousers": {
        "Package Contains": "1 Trouser",
        "Type": "Casual Trousers",
        "Length": "Regular"
    },
    "t-shirts": {
        "Package Contains": "1 T-Shirt",
        "Type": "T-Shirt",
        "Length": "Regular"
    },
    "shirts": {
        "Package Contains": "1 Shirt",
        "Type": "Casual Shirt",
        "Length": "Regular"
    },
    "dresses": {
        "Package Contains": "1 Dress",
        "Type": "A-Line Dress",
        "Length": "Knee Length"
    },
    "track pants": {
        "Package Contains": "1 Track Pant",
        "Type": "Track Pants",
        "Length": "Regular"
    },
    "clothing set": {
        "Package Contains": "1 T-Shirt, 1 Shorts",
        "Type": "Clothing Set",
        "Length": "Regular"
    },
    "clothing sets": {
        "Package Contains": "1 T-Shirt, 1 Shorts",
        "Type": "Clothing Set",
        "Length": "Regular"
    },
    "rompers": {
        "Package Contains": "1 Romper",
        "Type": "Romper",
        "Length": "Short"
    },
    "romper": {
        "Package Contains": "1 Romper",
        "Type": "Romper",
        "Length": "Short"
    }
}

MYNTRA_ARTICLE_TYPES = sorted(list(set([
    "T-Shirts", "Tshirts", "Shirts", "Tops", "Tunics", "Kurtas", "Kurta Sets", "Jackets", "Sweatshirts", 
    "Sweaters", "Cardigans", "Coats", "Shrugs", "Blazer", "Waistcoat", "Suits", "Nehru Jackets", 
    "Ponchos", "Raincoats", "Windcheaters", "Dresses", "Jumpsuits", "Rompers", "Romper", "Dungarees", "Sherwanis",
    "Trousers", "Shorts", "Jeans", "Track Pants", "Joggers", "Leggings", "Jeggings", "Capris", "Palazzos", 
    "Skirts", "Salwars", "Churidars", "Patialas", "Harem Pants", "Cargo Pants", "Tights", "Stockings",
    "Casual Shoes", "Sports Shoes", "Formal Shoes", "Sneakers", "Loafers", "Boots", "Sandals", "Slippers", 
    "Flip Flops", "Heels", "Flats", "Wedges", "Espadrilles", "Mules", "Clogs", "School Shoes",
    "Bras", "Briefs", "Trunks", "Boxers", "Camisoles", "Vests", "Thermal Tops", "Thermal Bottoms", 
    "Nightdress", "Nightsuits", "Bathrobes", "Pyjamas", "Shapewear", "Swimwear", "Board Shorts",
    "Lehenga Cholis", "Sarees", "Blouses", "Dupattas", "Kurtis", "Dhotis", "Gowns", "Anarkalis",
    "Watches", "Belts", "Wallets", "Sunglasses", "Caps", "Hats", "Socks", "Ties", "Pocket Squares", 
    "Mufflers", "Scarves", "Stoles", "Gloves", "Handbags", "Backpacks", "Duffle Bags", "Trolley Bags", 
    "Messenger Bags", "Clutches", "Purses", "Tote Bags", "Sling Bags", "Rucksacks", "Travel Pouches", 
    "Suitcases", "Laptop Bags", "Briefcases", "Keychains", "Umbrellas", "Hair Bands", "Hair Clips", 
    "Hair Pins", "Hair Accessories", "Headbands", "Ear Muffs", "Ties and Pocket Square Sets",
    "Necklaces", "Earrings", "Rings", "Bracelets", "Bangles", "Anklets", "Brooches", "Cufflinks", 
    "Nose Pins", "Pendants", "Jewellery Sets", "Mangalsutras",
    "Perfume", "Deodorant", "Lip Balm", "Lipstick", "Nail Polish", "Foundation", "Eyeliner", "Mascara", 
    "Eye Shadow", "Compact", "Blush", "Kajal", "Concealer", "Primer", "Makeup Brushes", "Hair Dryer", 
    "Hair Straightener", "Trimmers", "Shavers", "Epilators", "Face Wash", "Moisturizer", "Sunscreen", 
    "Shampoo", "Conditioner", "Hair Oil", "Body Wash", "Soap", "Body Lotion",
    "Clothing Set", "Clothing Sets"
])))

def get_singular_category(category_name):
    cat_lower = category_name.lower().strip()
    if cat_lower == "shorts":
        return "Short"
    elif cat_lower == "trousers":
        return "Trouser"
    elif cat_lower in ["t-shirts", "tshirts", "t-shirt", "tshirt"]:
        return "T-Shirt"
    elif cat_lower in ["shirts", "shirt"]:
        return "Shirt"
    elif cat_lower in ["dresses", "dress"]:
        return "Dress"
    elif cat_lower in ["track pants", "track pant"]:
        return "Track Pant"
    elif cat_lower.endswith("s") and len(cat_lower) > 1:
        return category_name[:-1]
    return category_name

def parse_fabric_and_type(row, selected_cat=""):
    fabric_val = ""
    fabric_type_val = ""
    
    # 1. Gather all text info from row
    material_text = str(row.get('MATERIAL', '')).strip().upper()
    fabric_text = str(row.get('FABRIC', '')).strip().upper()
    combined_text = f"{material_text} {fabric_text}"
    
    # Skip placeholder values
    if "(NIL)" in combined_text:
        combined_text = combined_text.replace("(NIL)", "")
    
    # 2. Extract Fabric
    if "COTTON" in combined_text:
        fabric_val = "Cotton"
    elif "POLYESTER" in combined_text or "PES" in combined_text:
        fabric_val = "Polyester"
    elif "LINEN" in combined_text:
        fabric_val = "Linen"
    elif "NYLON" in combined_text:
        fabric_val = "Nylon"
    elif "RAYON" in combined_text or "VISCOSE" in combined_text:
        fabric_val = "Viscose Rayon"
    elif "SILK" in combined_text:
        fabric_val = "Silk"
    elif "WOOL" in combined_text:
        fabric_val = "Wool"
    elif "ACRYLIC" in combined_text:
        fabric_val = "Acrylic"
    elif "MODAL" in combined_text:
        fabric_val = "Modal"
    else:
        # Fallback: if we have a non-nil fabric/material text that isn't too long, use it. Otherwise "Cotton"
        words = [w.capitalize() for w in fabric_text.split() if w and w not in ["(NIL)", "NIL"]]
        if words and len(words) <= 3:
            fabric_val = " ".join(words)
        else:
            fabric_val = "Cotton"
            
    # 3. Extract Fabric Type
    if "JERSEY" in combined_text:
        fabric_type_val = "Jersey"
    elif "DENIM" in combined_text:
        fabric_type_val = "Denim"
    elif "FLEECE" in combined_text:
        fabric_type_val = "Fleece"
    elif "FLANNEL" in combined_text:
        fabric_type_val = "Flannel"
    elif "TERRY" in combined_text:
        fabric_type_val = "French Terry"
    elif "VELOUR" in combined_text:
        fabric_type_val = "Velour"
    elif "RIB" in combined_text:
        fabric_type_val = "Ribbed"
    elif "PIQUE" in combined_text:
        fabric_type_val = "Pique"
    elif "WAFFLE" in combined_text:
        fabric_type_val = "Waffle"
    elif "WOVEN" in combined_text:
        fabric_type_val = "Woven"
    elif "KNITTED" in combined_text or "KNIT" in combined_text:
        fabric_type_val = "Knitted"
    else:
        # Default fallback based on category
        cat_lower = selected_cat.lower()
        if "t-shirt" in cat_lower or "tshirt" in cat_lower:
            fabric_type_val = "Jersey"
        else:
            fabric_type_val = "Other"
            
    return fabric_val, fabric_type_val

def get_closest_match(val, choices):
    if not choices:
        return str(val).strip()
    
    val_str = str(val).strip()
    val_lower = val_str.lower()
    
    # 1. Exact match (case insensitive)
    for ch in choices:
        if ch.lower() == val_lower:
            return ch
            
    # 2. Substring match or keyword overlap
    best_ch = None
    best_score_ch = -1
    for ch in choices:
        ch_lower = ch.lower()
        # Find number of common words or characters
        words_val = set(val_lower.replace('-', ' ').replace('_', ' ').split())
        words_ch = set(ch_lower.replace('-', ' ').replace('_', ' ').split())
        overlap = len(words_val.intersection(words_ch))
        
        # Substring boost
        if val_lower in ch_lower or ch_lower in val_lower:
            overlap += 2
            
        if overlap > best_score_ch:
            best_score_ch = overlap
            best_ch = ch
            
    if best_ch and best_score_ch > 0:
        return best_ch
        
    # 3. Fallback: Edit distance
    def edit_distance(s1, s2):
        if len(s1) > len(s2):
            s1, s2 = s2, s1
        distances = range(len(s1) + 1)
        for i2, c2 in enumerate(s2):
            distances_ = [i2+1]
            for i1, c1 in enumerate(s1):
                if c1 == c2:
                    distances_.append(distances[i1])
                else:
                    distances_.append(1 + min((distances[i1], distances[i1 + 1], distances_[-1])))
            distances = distances_
        return distances[-1]
        
    best_dist = 9999
    best_ch_dist = choices[0]
    for ch in choices:
        dist = edit_distance(val_lower, ch.lower())
        if dist < best_dist:
            best_dist = dist
            best_ch_dist = ch
            
    return best_ch_dist

def run_automation_core(params, dry_run=False):
    global current_task_status
    run_id = params.get('run_id', '')
    current_task_status = {
        "status": "running",
        "progress": 0.0,
        "current_step": "Initializing",
        "error": None,
        "run_id": run_id
    }
    
    try:
        log_buffer.log("INFO", "=== STARTING AUTOMATION TASK ===")
        log_buffer.log("INFO", f"Task Type: {'DRY RUN ANALYSIS' if dry_run else 'COMPLETE SKU GENERATION'}")
        
        # 1. Extract paths and variables
        item_path = params.get('item_path')
        content_path = params.get('content_path')
        template_path = params.get('template_path')
        output_dir = params.get('output_dir')
        
        selected_cat = params.get('category', 'Shorts')
        isp_rule = params.get('isp_rule', 'Discounted')
        isp_discount = float(params.get('isp_discount', 50.0))
        raw_products_text = params.get('products_filter', '').strip()
        
        use_vision = params.get('use_vision', False)
        product_images_dir = params.get('product_images_dir', '').strip()
        vision_cache = {}
        
        if use_vision:
            if not product_images_dir or not os.path.exists(product_images_dir):
                log_buffer.log("WARNING", "AI Vision auto-fill was enabled but the Product Images Folder was not found or is empty. Falling back to default rules.")
                use_vision = False
            else:
                log_buffer.log("INFO", f"AI Vision enabled! Scanning '{product_images_dir}' for product photos...")
        
        # Validation of paths
        paths = [("Item Directory", item_path), ("Fresh Template", template_path)]
            
        for name, path in paths:
            if not path or not os.path.exists(path):
                log_buffer.log("ERROR", f"{name} not found at path: {path}")
                raise FileNotFoundError(f"{name} file could not be found.")
                
        # Validate content file if provided
        if content_path and str(content_path).strip():
            if not os.path.exists(content_path):
                log_buffer.log("ERROR", f"Content File not found at path: {content_path}")
                raise FileNotFoundError("Content File could not be found.")
        
        current_task_status["progress"] = 0.10
        current_task_status["current_step"] = "Loading Item Directory"
        
        # 2. Load Item Directory
        log_buffer.log("INFO", "Loading Item Directory (Master Sheet) into memory...")
        try:
            xl = pd.ExcelFile(item_path)
            sheet_names = xl.sheet_names
            selected_sheet = None
            for s in sheet_names:
                if 'report' in s.lower():
                    selected_sheet = s
                    break
            if not selected_sheet:
                selected_sheet = sheet_names[0]
                log_buffer.log("WARNING", f"Worksheet named 'Report' not found. Falling back to first sheet: '{selected_sheet}'")
            item_df = pd.read_excel(item_path, sheet_name=selected_sheet)
        except Exception as sheet_err:
            log_buffer.log("ERROR", f"Failed to load Item Directory sheets: {str(sheet_err)}")
            raise ValueError(f"Could not read the uploaded Item Directory: {str(sheet_err)}")

        log_buffer.log("SUCCESS", f"Loaded Item Directory: {len(item_df)} items (Sheet: '{selected_sheet}').")
        
        current_task_status["progress"] = 0.25
        current_task_status["current_step"] = "Loading Content File"
        
        # 3. Load Content Sheet
        if content_path and str(content_path).strip():
            log_buffer.log("INFO", f"Loading Content File from: {content_path}...")
            try:
                c_xl = pd.ExcelFile(content_path)
                c_sheet_names = c_xl.sheet_names
                selected_c_sheet = None
                for s in c_sheet_names:
                    clean_s = s.lower().replace(" ", "").replace("_", "").replace("-", "")
                    if 'marketplaced2c' in clean_s:
                        selected_c_sheet = s
                        break
                if not selected_c_sheet:
                    selected_c_sheet = c_sheet_names[0]
                    log_buffer.log("WARNING", f"Worksheet named 'MarketplaceD2C' not found in Content File. Falling back to first sheet: '{selected_c_sheet}'")
                content_df = pd.read_excel(content_path, sheet_name=selected_c_sheet)
            except Exception as c_sheet_err:
                log_buffer.log("ERROR", f"Failed to load Content File sheets: {str(c_sheet_err)}")
                raise ValueError(f"Could not read the uploaded Content File: {str(c_sheet_err)}")
            log_buffer.log("SUCCESS", f"Loaded Content File: {len(content_df)} catalog templates (Sheet: '{selected_c_sheet}').")
        else:
            log_buffer.log("INFO", "No Content File provided. Skipping description matching and using defaults.")
            content_df = pd.DataFrame(columns=["Item Name", "Myntra Title (List View Name)", "PDP Name (Myntra)", "Description", "SKU"])
        
        current_task_status["progress"] = 0.40
        current_task_status["current_step"] = "Loading Learning Rules Model"
        
        # 4. Reconcile rules, brand rules, and sizing charts dynamically from myntra_rules_db.json
        sizing_measurements_db = {}
        fabric_static_db = {}
        
        # Load user-editable sizechart mappings database dynamically on run
        sizechart_mappings = {}
        mappings_file = "myntra_sizechart_mappings.json"
        if os.path.exists(mappings_file):
            try:
                with open(mappings_file, "r", encoding="utf-8") as f:
                    sizechart_mappings = json.load(f)
                log_buffer.log("SUCCESS", f"Loaded custom sizechart mappings database from '{mappings_file}'.")
            except Exception as e:
                log_buffer.log("WARNING", f"Failed to load sizechart mappings file: {str(e)}")
        
        FOOTWEAR_CATEGORIES = [
            "ballerinas", "casual shoes", "sports shoes", "formal shoes", "sneakers", "loafers", "boots", 
            "sandals", "slippers", "flip flops", "heels", "flats", "wedges", "espadrilles", "mules", "clogs"
        ]
        is_footwear = selected_cat.lower().strip() in FOOTWEAR_CATEGORIES
        
        if is_footwear:
            log_buffer.log("INFO", f"Footwear category identified ('{selected_cat}'). Applying brand 'toothless' and Footwear defaults...")
            fabric_static_db = rules_db.get("footwear_defaults", {})
            sizing_measurements_db = rules_db.get("footwear_sizecharts", {}).get("FOOTWEAR", {})
            log_buffer.log("SUCCESS", f"Loaded {len(sizing_measurements_db)} footwear sizes from learning model.")
        else:
            log_buffer.log("INFO", f"Apparel category identified ('{selected_cat}'). Applying brand 'Purple United Kids' and Apparel defaults...")
            fabric_static_db = rules_db.get("apparel_defaults", {})
            
            apparel_charts = rules_db.get("apparel_sizecharts", {})
            
            # Map selected_cat to sizechart_category from mappings JSON
            mapped_sizechart_cat = None
            for key_cat, val_cat in sizechart_mappings.items():
                if key_cat.lower().strip() == selected_cat.lower().strip():
                    mapped_sizechart_cat = val_cat.get("sizechart_category")
                    break
            
            if mapped_sizechart_cat:
                sizechart_cat = mapped_sizechart_cat
            else:
                sizechart_cat = None
                for key in apparel_charts.keys():
                    if selected_cat.upper().strip() in key or key in selected_cat.upper().strip():
                        sizechart_cat = key
                        break
            
            if not sizechart_cat:
                cat_lower = selected_cat.lower()
                if "top" in cat_lower or "t-shirt" in cat_lower or "tshirt" in cat_lower or "shirt" in cat_lower:
                    sizechart_cat = "BOYS T-SHIRT"
                elif "shorts" in cat_lower:
                    sizechart_cat = "BOYS SHORTS"
                elif "trousers" in cat_lower or "jogger" in cat_lower or "pants" in cat_lower or "track" in cat_lower:
                    sizechart_cat = "BOYS LOWER /JOGGER"
                elif "romper" in cat_lower:
                    sizechart_cat = "ROMPERS"
                elif "jacket" in cat_lower:
                    sizechart_cat = "JACKET"
                elif "sweatshirt" in cat_lower:
                    sizechart_cat = "SWEATSHIRT"
                else:
                    sizechart_cat = list(apparel_charts.keys())[0] if apparel_charts else None
            
            if isinstance(sizechart_cat, dict):
                default_chart_name = sizechart_cat.get("DEFAULT", "")
                sizing_measurements_db = apparel_charts.get(default_chart_name, {})
                log_buffer.log("SUCCESS", f"Loaded custom multi-gender size chart mappings. Default fallback: '{default_chart_name}' ({len(sizing_measurements_db)} sizes).")
            else:
                sizing_measurements_db = apparel_charts.get(sizechart_cat, {}) if sizechart_cat else {}
                log_buffer.log("SUCCESS", f"Loaded '{sizechart_cat}' size chart containing {len(sizing_measurements_db)} sizes from learning model.")
        
        current_task_status["progress"] = 0.55
        current_task_status["current_step"] = "Matching Target Template"
        
        # 5. Dynamically load target Template Excel
        log_buffer.log("INFO", "Opening fresh Target SKU Template to prepare for injection...")
        wb = openpyxl.load_workbook(template_path)
        
        category_sheets = [s for s in wb.sheetnames if s not in ["__INSTRUCTIONS", "masterdata"]]
        if not category_sheets:
            raise ValueError("No product category sheets found in fresh SKU template! Must contain a category sheet.")
        
        target_sheet_name = None
        for sheet_name in wb.sheetnames:
            if sheet_name not in ["__INSTRUCTIONS", "masterdata"]:
                if sheet_name.lower().strip() == selected_cat.lower().strip():
                    target_sheet_name = sheet_name
                    break
        
        if not target_sheet_name:
            target_sheet_name = category_sheets[0]
            log_buffer.log("WARNING", f"Could not find a sheet named '{selected_cat}'. Falling back to first available sheet '{target_sheet_name}'.")
        
        sheet = wb[target_sheet_name]
        log_buffer.log("SUCCESS", f"Target Sheet identified: '{target_sheet_name}' ({sheet.max_row} rows present).")
        
        tech_keys_row = [cell.value for cell in sheet[3]]
        col_mapping = {}
        for col_idx, key in enumerate(tech_keys_row, start=1):
            if key:
                col_mapping[str(key).strip()] = col_idx
                
        log_buffer.log("INFO", f"Parsed {len(col_mapping)} tech columns from fresh target template.")
        
        # Extract masterdata options for season, closure, and type
        masterdata_choices = {
            "season": [],
            "closure": [],
            "type": []
        }
        
        if "masterdata" in wb.sheetnames:
            try:
                log_buffer.log("INFO", "Scanning 'masterdata' sheet for dropdown validation lists...")
                m_sheet = wb["masterdata"]
                
                # Define signatures
                season_sigs = {'summer', 'winter', 'spring', 'fall', 'spring-summer', 'fall-winter', 'autumn'}
                closure_sigs = {'tie-ups', 'zip', 'button', 'hook and eye', 'lace-ups', 'buckle', 'velcro', 'slip-on', 'backstrap', 'drawstring', 'elasticated'}
                type_sigs = {'regular', 'casual', 't-shirt', 'shirt', 'dress', 'trousers', 'shorts', 'blouson', 'wrap', 'boxy', 'peplum', 'a-line', 'mules', 'ballerinas', 'sneakers'}
                
                # Scan each column in masterdata sheet
                for col_idx in range(1, m_sheet.max_column + 1):
                    header_val = m_sheet.cell(row=1, column=col_idx).value
                    if header_val is None:
                        continue
                    header_str = str(header_val).strip().lower()
                    
                    # Skip generic metadata columns for closure and type to avoid contamination
                    is_generic_meta = any(x in header_str for x in ['brand', 'country', 'origin', 'color', 'colour', 'agegroup', 'usage', 'fashiontype', 'year', 'articletype'])
                    
                    # Sample values to determine the column's semantic type
                    vals = []
                    for r in range(2, min(m_sheet.max_row + 1, 500)):
                        v = m_sheet.cell(row=r, column=col_idx).value
                        if v is not None:
                            vals.append(str(v).strip().lower())
                    
                    unique_vals = set(vals)
                    if not unique_vals:
                        continue
                        
                    # Calculate matching scores
                    season_score = sum(1 for v in unique_vals if v in season_sigs)
                    if header_str == 'season':
                        season_score += 100
                        
                    closure_score = sum(1 for v in unique_vals if v in closure_sigs)
                    if 'closure' in header_str or 'fastening' in header_str:
                        closure_score += 100
                        
                    type_score = sum(1 for v in unique_vals if v in type_sigs)
                    if 'type' in header_str and header_str not in ['articletype', 'fabric type', 'fashiontype'] and not is_generic_meta:
                        type_score += 50
                    elif is_generic_meta:
                        type_score = 0
                        
                    # Extract column choices
                    def get_column_choices(c_idx):
                        choices = []
                        for r in range(2, m_sheet.max_row + 1):
                            v = m_sheet.cell(row=r, column=c_idx).value
                            if v is not None:
                                vs = str(v).strip()
                                if vs and vs not in choices:
                                    choices.append(vs)
                        return choices
                        
                    if season_score > 0 and season_score > masterdata_choices.get('_season_score', -1):
                        masterdata_choices['season'] = get_column_choices(col_idx)
                        masterdata_choices['_season_score'] = season_score
                        
                    if closure_score > 0 and closure_score > masterdata_choices.get('_closure_score', -1):
                        masterdata_choices['closure'] = get_column_choices(col_idx)
                        masterdata_choices['_closure_score'] = closure_score
                        
                    if type_score > 0 and type_score > masterdata_choices.get('_type_score', -1):
                        masterdata_choices['type'] = get_column_choices(col_idx)
                        masterdata_choices['_type_score'] = type_score
                        
                log_buffer.log("SUCCESS", f"Identified masterdata validations: Season ({len(masterdata_choices['season'])} options), Closure ({len(masterdata_choices['closure'])} options), Type ({len(masterdata_choices['type'])} options).")
            except Exception as ex:
                log_buffer.log("WARNING", f"Failed to parse masterdata sheet validations: {str(ex)}. Using fallback mapping rules.")
        
        current_task_status["progress"] = 0.70
        current_task_status["current_step"] = "Processing Matches"
        
        # 6. Group the Item Directory rows by ITEM NAME
        log_buffer.log("INFO", "Grouping size-specific rows in Item Directory by Style Code ('ITEM NAME')...")
        has_content_file = bool(content_path and str(content_path).strip())
        
        if has_content_file:
            content_styles = set(content_df['Item Name'].dropna().astype(str).str.strip().str.upper())
            matching_items = item_df[item_df['ITEM NAME'].astype(str).str.strip().str.upper().isin(content_styles)]
            log_buffer.log("INFO", f"Total size rows in Item Directory: {len(item_df)}")
            log_buffer.log("SUCCESS", f"Filtered size rows matching the Content Sheet: {len(matching_items)} rows.")
            
            if len(matching_items) == 0:
                log_buffer.log("WARNING", "Zero overlap detected between Item Directory['ITEM NAME'] and Content File['Item Name']!")
                item_df['clean_name'] = item_df['ITEM NAME'].astype(str).str.strip().str.upper()
                content_df['clean_name'] = content_df['Item Name'].astype(str).str.strip().str.upper()
                matching_items = item_df[item_df['clean_name'].isin(content_df['clean_name'])]
                log_buffer.log("SUCCESS", f"Soft matching yielded: {len(matching_items)} rows.")
                
            if len(matching_items) == 0:
                raise ValueError("Aborting process. No matching products found between your Item Directory and your Content File.")
        else:
            matching_items = item_df.copy()
            log_buffer.log("INFO", f"Total size rows in Item Directory (no Content File filter): {len(matching_items)} rows.")
            
        grouped_styles = matching_items.groupby('ITEM NAME')
        
        # Filters to apply
        product_lines = [line.strip().upper() for line in raw_products_text.splitlines() if line.strip()]
        has_filter = False
        specific_items_set = set()
        if product_lines and product_lines != ["ALL"]:
            has_filter = True
            specific_items_set = set(product_lines)
            log_buffer.log("INFO", f"Listing only {len(specific_items_set)} specific product(s).")
        else:
            log_buffer.log("INFO", "No filter applied — listing ALL products.")
            
        # 7. Merge & Process Rows
        rows_to_insert = []
        
        content_df_lookup = content_df.copy()
        content_df_lookup['_style_upper'] = content_df_lookup['Item Name'].astype(str).str.strip().str.upper()
        
        content_color_col = None
        for col_name in ['Color', 'COLOR', 'Colour', 'COLOUR', 'Item Color', 'Brand Color', 'Brand Colour', 'SHADE NAME', 'Shade Name', 'Shade name']:
            if col_name in content_df_lookup.columns:
                content_color_col = col_name
                break
        
        if content_color_col:
            content_df_lookup['_color_upper'] = content_df_lookup[content_color_col].astype(str).str.strip().str.upper()
            log_buffer.log("INFO", f"Content file color column detected: '{content_color_col}'.")
        else:
            content_df_lookup['_color_upper'] = ''
            log_buffer.log("WARNING", "No color column found in content file.")
        
        content_sku_key = {}
        content_dual_key = {}
        content_style_key = {}
        for _, crow in content_df_lookup.iterrows():
            sk = crow['_style_upper']
            ck = crow['_color_upper']
            sku_val = str(crow.get('SKU', '')).strip().upper()
            
            if sku_val:
                content_sku_key[sku_val] = crow
                
            dual_k = f"{sk}||{ck}"
            if dual_k not in content_dual_key:
                content_dual_key[dual_k] = crow
            if sk not in content_style_key:
                content_style_key[sk] = crow
        
        inserted_count = 0
        
        for style_code, group in grouped_styles:
            style_code_upper = str(style_code).strip().upper()
            
            # Verify if we have matching SKU or Style descriptive content if content file is provided
            if has_content_file:
                has_content_for_group = False
                for _, r in group.iterrows():
                    b_color = str(r['COLOR']).strip() if pd.notna(r['COLOR']) else (str(r['Item Color']).strip() if pd.notna(r['Item Color']) else "")
                    van = f"{style_code}-{b_color.upper()}"
                    if van in content_sku_key or style_code_upper in content_style_key:
                        has_content_for_group = True
                        break
                        
                if not has_content_for_group:
                    log_buffer.log("WARNING", f"Descriptive content for style '{style_code}' not found in Content File. Skipping.")
                    continue
            
            for idx, row in group.iterrows():
                brand_color = str(row['COLOR']).strip() if pd.notna(row['COLOR']) else (str(row['Item Color']).strip() if pd.notna(row['Item Color']) else "")
                item_color_key = f"{style_code}-{brand_color.upper()}"
                
                if has_filter and item_color_key not in specific_items_set:
                    continue
                
                vendor_article_num = f"{style_code}-{brand_color.upper()}"
                
                # AI Vision classification lookup
                vision_preds = None
                if use_vision:
                    van_upper = vendor_article_num.upper()
                    if van_upper not in vision_cache:
                        try:
                            all_files = os.listdir(product_images_dir)
                            matching_files = [f for f in all_files if f.upper().startswith(van_upper)]
                            matching_imgs = [f for f in matching_files if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
                            matching_imgs.sort()
                            
                            if len(matching_imgs) >= 2:
                                front_img_path = os.path.join(product_images_dir, matching_imgs[0])
                                back_img_path = os.path.join(product_images_dir, matching_imgs[1])
                                
                                log_buffer.log("INFO", f"[AI VISION] Found photos for '{vendor_article_num}': {matching_imgs[0]}, {matching_imgs[1]}")
                                log_buffer.log("INFO", f"[AI VISION] Analyzing product visuals via Gemini Vision...")
                                
                                # Call analyzer
                                res = vision_classifier.analyze_product_images(front_img_path, back_img_path, selected_cat, template_path)
                                if "predictions" in res:
                                    vision_cache[van_upper] = res["predictions"]
                                    log_buffer.log("SUCCESS", f"[AI VISION] Classified attributes: {json.dumps(res['predictions'])}")
                                    
                                    # Automatically register prediction to learnings database for review/correction
                                    try:
                                        vision_classifier.save_learning(
                                            vendor_article_num=vendor_article_num,
                                            category=selected_cat,
                                            front_img_path=front_img_path,
                                            back_img_path=back_img_path,
                                            labels=res["predictions"]
                                        )
                                        log_buffer.log("INFO", f"[AI VISION] Saved predictions for '{vendor_article_num}' to learnings database.")
                                    except Exception as db_err:
                                        log_buffer.log("WARNING", f"[AI VISION] Failed to save prediction to learnings database: {str(db_err)}")
                                else:
                                    log_buffer.log("WARNING", f"[AI VISION] Failed to analyze: {res.get('error', 'Unknown error')}")
                                    vision_cache[van_upper] = None
                            elif len(matching_imgs) == 1:
                                log_buffer.log("WARNING", f"[AI VISION] Found only 1 photo for '{vendor_article_num}', but front and back are required. Skipping AI.")
                                vision_cache[van_upper] = None
                            else:
                                vision_cache[van_upper] = None
                        except Exception as ex:
                            log_buffer.log("WARNING", f"[AI VISION] Error scanning files or analyzing: {str(ex)}")
                            vision_cache[van_upper] = None
                            
                    vision_preds = vision_cache.get(van_upper)
                
                # Direct SKU-based lookup matching, falling back to dual/style-based match
                content_row = None
                if vendor_article_num in content_sku_key:
                    content_row = content_sku_key[vendor_article_num]
                else:
                    dual_k = f"{style_code_upper}||{brand_color.upper()}"
                    if dual_k in content_dual_key:
                        content_row = content_dual_key[dual_k]
                    else:
                        content_row = content_style_key.get(style_code_upper, None)
                
                if content_row is not None:
                    myntra_title = content_row.get("Myntra Title (List View Name)", "")
                    pdp_name = content_row.get("PDP Name (Myntra)", "")
                    pdp_description = content_row.get("Description", "")
                else:
                    myntra_title = ""
                    pdp_name = f"{row.get('Brand', '')} {style_code} {selected_cat}"
                    pdp_description = ""
                    
                sku_data = {}
                barcode = str(row['ITEM CODE']).strip()
                item_size = str(row['SIZE']).strip()
                mrp = row['MRP']
                
                if isp_rule == "Same as MRP":
                    isp = mrp
                else:
                    isp = round(mrp * (1.0 - (isp_discount / 100.0)))
                
                sku_data['vendorSkuCode'] = barcode
                sku_data['GTIN'] = "" # Kept blank as requested
                sku_data['SKUCode'] = barcode
                sku_data['MRP'] = mrp
                sku_data['ISP'] = isp
                
                sku_data['Brand Size'] = item_size
                sku_data['Standard Size'] = item_size
                sku_data['is Standard Size present on Label'] = "Yes"
                
                sku_data['Brand Colour (Remarks)'] = brand_color
                sku_data['Prominent Colour'] = brand_color.capitalize()
                
                sku_data['vendorArticleNumber'] = f"{style_code}-{brand_color.upper()}"
                sku_data['vendorArticleName'] = pdp_name if pd.notna(pdp_name) else (myntra_title if pd.notna(myntra_title) else f"{row.get('Brand', '')} {style_code} {selected_cat}")
                sku_data['List View Name'] = myntra_title if pd.notna(myntra_title) else sku_data['vendorArticleName']
                sku_data['Product Details'] = pdp_description if pd.notna(pdp_description) else ""
                
                # Hardcoded Brand name override based on company product line rules
                sku_data['brand'] = "toothless" if is_footwear else "Purple United Kids"
                sku_data['HSN'] = str(row.get('HS CODE', '')).replace('.0', '').strip()
                
                imported_domestic = str(row.get('IMPORTED/DOMESTIC', '')).strip().lower()
                if 'import' in imported_domestic:
                    sku_data['Country Of Origin'] = "Vietnam"
                else:
                    sku_data['Country Of Origin'] = "India"
                
                gender = str(row.get('GENDER', '')).strip().upper()
                age_group = str(row.get('AGE GROUP', '')).strip().upper()
                
                if 'GIRL' in gender and 'KID' in age_group:
                    sku_data['AgeGroup'] = "Kids-Girls"
                elif 'BOY' in gender and 'KID' in age_group:
                    sku_data['AgeGroup'] = "Kids-Boys"
                elif 'GIRL' in gender:
                    sku_data['AgeGroup'] = "Girls"
                elif 'BOY' in gender:
                    sku_data['AgeGroup'] = "Boys"
                elif 'WOMEN' in gender or 'FEMALE' in gender:
                    sku_data['AgeGroup'] = "Women"
                else:
                    sku_data['AgeGroup'] = "Men"
                    
                sku_data['articleType'] = target_sheet_name
                sku_data['FashionType'] = "Fashion"
                sku_data['Usage'] = "Casual"
                sku_data['Year'] = 2026
                
                # Strict Myntra Dropdown Season matching
                season = str(row.get('SEASON', '')).strip().upper()
                if 'WINTER' in season or 'FALL' in season or 'AUTUMN' in season:
                    sku_data['season'] = "Fall-Winter"
                else:
                    sku_data['season'] = "Spring-Summer"
                
                # Normalized offline size database lookup with dynamic column mapping
                norm_size = normalize_size_label(item_size)
                
                # Resolve custom sizechart mapping entry
                selected_entry = None
                selected_mappings = {}
                for cat_k, cat_v in sizechart_mappings.items():
                    if cat_k.lower().strip() == selected_cat.lower().strip():
                        selected_entry = cat_v
                        selected_mappings = cat_v.get("mappings", {})
                        break
                
                # Resolve the sizechart category dynamically based on gender
                resolved_measurements = None
                resolved_chart_name = None
                
                if selected_entry:
                    sizechart_cat_config = selected_entry.get("sizechart_category")
                    if isinstance(sizechart_cat_config, dict):
                        # Determine preferred and fallback categories
                        preferred_gender = "GIRLS" if "GIRL" in gender else ("BOYS" if "BOY" in gender else "DEFAULT")
                        fallback_gender = "BOYS" if preferred_gender == "GIRLS" else ("GIRLS" if preferred_gender == "BOYS" else None)
                        
                        preferred_chart = sizechart_cat_config.get(preferred_gender)
                        if preferred_chart:
                            preferred_chart = preferred_chart.upper().strip()
                        fallback_chart = sizechart_cat_config.get(fallback_gender) if fallback_gender else None
                        if fallback_chart:
                            fallback_chart = fallback_chart.upper().strip()
                        default_chart = sizechart_cat_config.get("DEFAULT")
                        if default_chart:
                            default_chart = default_chart.upper().strip()
                        
                        # 1. Try preferred gender
                        if preferred_chart and preferred_chart in apparel_charts:
                            if norm_size in apparel_charts[preferred_chart]:
                                resolved_measurements = apparel_charts[preferred_chart][norm_size]
                                resolved_chart_name = preferred_chart
                                
                        # 2. Try fallback gender (opposite)
                        if not resolved_measurements and fallback_chart and fallback_chart in apparel_charts:
                            if norm_size in apparel_charts[fallback_chart]:
                                resolved_measurements = apparel_charts[fallback_chart][norm_size]
                                resolved_chart_name = fallback_chart
                                
                        # 3. Try default
                        if not resolved_measurements and default_chart and default_chart in apparel_charts:
                            if norm_size in apparel_charts[default_chart]:
                                resolved_measurements = apparel_charts[default_chart][norm_size]
                                resolved_chart_name = default_chart
                                
                    elif isinstance(sizechart_cat_config, str):
                        # Simple string mapping
                        chart_upper = sizechart_cat_config.upper().strip()
                        if chart_upper in apparel_charts:
                            if norm_size in apparel_charts[chart_upper]:
                                resolved_measurements = apparel_charts[chart_upper][norm_size]
                                resolved_chart_name = chart_upper
                                
                # If not resolved via mappings, fall back to global sizing_measurements_db
                if not resolved_measurements and norm_size in sizing_measurements_db:
                    resolved_measurements = sizing_measurements_db[norm_size]
                    resolved_chart_name = sizechart_cat.get("DEFAULT") if isinstance(sizechart_cat, dict) else sizechart_cat
                    
                if resolved_measurements:
                    for m_col, m_val in resolved_measurements.items():
                        # Translate point of measure name to actual technical column name
                        translated_col = selected_mappings.get(m_col, m_col)
                        sku_data[translated_col] = m_val
                
                # Enforce company hardcoded dropdown selections
                if not is_footwear:
                    sku_data['Fit'] = "Regular Fit"
                    sku_data['Brand Fit Name'] = "NA"
                    sku_data['Waist Rise'] = "Mid-Rise"
                    sku_data['Multipack Set'] = ""  # Remains blank for clothing/apparel
                else:
                    sku_data['Multipack Set'] = "NA" # Set to NA for footwear
                
                selected_cat_key = selected_cat.lower().strip()
                singular_cat = get_singular_category(selected_cat)
                
                other_keywords = ["shorts", "short", "trousers", "trouser", "t-shirt", "tshirts", "tshirt", "shirt", "dress", "dresses", "pant", "pants"]
                current_cat_keywords = {singular_cat.lower(), selected_cat_key}
                if "t-shirt" in current_cat_keywords or "t-shirts" in current_cat_keywords:
                    current_cat_keywords.update(["t-shirt", "tshirts", "tshirt", "t-shirts"])
                check_keywords = [w for w in other_keywords if w not in current_cat_keywords]

                for s_col, s_val in fabric_static_db.items():
                    if s_col not in sku_data or pd.isna(sku_data[s_col]) or sku_data[s_col] == "":
                        val_str = str(s_val).lower().strip()
                        
                        is_mismatch = False
                        if s_col in ["Package Contains", "Type", "Length"]:
                            if any(kw in val_str for kw in check_keywords):
                                is_mismatch = True
                            if s_col == "Length":
                                if selected_cat_key in ["trousers", "t-shirts", "shirts", "track pants"]:
                                    if any(kw in val_str for kw in ["knee", "midi", "maxi", "thigh"]):
                                        is_mismatch = True
                                elif selected_cat_key == "shorts":
                                    if any(kw in val_str for kw in ["full", "regular", "midi", "maxi"]):
                                        is_mismatch = True
                        
                        if is_mismatch:
                            if selected_cat_key in CATEGORY_DEFAULTS and s_col in CATEGORY_DEFAULTS[selected_cat_key]:
                                sku_data[s_col] = CATEGORY_DEFAULTS[selected_cat_key][s_col]
                            elif s_col == "Package Contains":
                                sku_data[s_col] = f"1 {singular_cat}"
                            else:
                                sku_data[s_col] = ""
                        else:
                            sku_data[s_col] = s_val
                            
                for special_col in ["Package Contains", "Type", "Length"]:
                    if special_col not in sku_data or pd.isna(sku_data[special_col]) or sku_data[special_col] == "":
                        if selected_cat_key in CATEGORY_DEFAULTS and special_col in CATEGORY_DEFAULTS[selected_cat_key]:
                            sku_data[special_col] = CATEGORY_DEFAULTS[selected_cat_key][special_col]
                        elif special_col == "Package Contains":
                            sku_data[special_col] = f"1 {singular_cat}"
                
                # --- Strict Formatting & Company Overrides ---
                sku_data['GTIN'] = "" # GTIN remains blank
                
                # Enforce dynamic Package Contains based on category
                selected_cat_key = selected_cat.lower().strip()
                singular_cat = get_singular_category(selected_cat)
                
                if selected_cat_key in CATEGORY_DEFAULTS and "Package Contains" in CATEGORY_DEFAULTS[selected_cat_key]:
                    sku_data["Package Contains"] = CATEGORY_DEFAULTS[selected_cat_key]["Package Contains"]
                else:
                    sku_data["Package Contains"] = f"1 {singular_cat.capitalize()}"
                
                # Enforce Type and Length if category defaults exist
                if selected_cat_key in CATEGORY_DEFAULTS:
                    for d_col, d_val in CATEGORY_DEFAULTS[selected_cat_key].items():
                        sku_data[d_col] = d_val
                
                if not is_footwear:
                    sku_data['Fit'] = "Regular Fit"
                    sku_data['Brand Fit Name'] = "NA"
                    sku_data['Waist Rise'] = "Mid-Rise"
                    sku_data['Weave Type'] = "Knitted"
                    sku_data['Lining'] = "NA"
                    
                    # Resolve Fabric and Fabric Type from Item Directory if missing
                    if 'Fabric' not in sku_data or pd.isna(sku_data['Fabric']) or sku_data['Fabric'] == "":
                        fabric_val, fabric_type_val = parse_fabric_and_type(row, selected_cat)
                        sku_data['Fabric'] = fabric_val
                    
                    sku_data['Fabric Type'] = sku_data.get('Fabric', '')
                    
                    # Knit or Woven: Summer -> Knit, Winter -> Woven
                    if sku_data.get('season') == "Fall-Winter":
                        sku_data['Knit or Woven'] = "Woven"
                    else:
                        sku_data['Knit or Woven'] = "Knit"
                            
                    # User overrides for Apparel
                     # Wash Care, Body or Garment Size, Number of Items, Net Quantity Unit, Net Quantity
                    sku_data['Wash Care'] = "Hand Wash"
                    sku_data['materialCareDescription'] = "Hand Wash"
                    if 'Body or Garment Size' not in sku_data or pd.isna(sku_data['Body or Garment Size']) or sku_data['Body or Garment Size'] == "":
                        sku_data['Body or Garment Size'] = "Garment Measurements in"
                    if 'Number of Items' not in sku_data or pd.isna(sku_data['Number of Items']) or sku_data['Number of Items'] == "":
                        sku_data['Number of Items'] = 1
                    if 'Net Quantity Unit' not in sku_data or pd.isna(sku_data['Net Quantity Unit']) or sku_data['Net Quantity Unit'] == "":
                        sku_data['Net Quantity Unit'] = "Piece"
                    if 'Net Quantity' not in sku_data or pd.isna(sku_data['Net Quantity']) or sku_data['Net Quantity'] == "":
                        sku_data['Net Quantity'] = 1
                else:
                    sku_data['materialCareDescription'] = "Wipe it with a clean dry cloth"
                
                if "clothing set" in selected_cat.lower().strip():
                    sku_data['Multipack Set'] = "2"
                else:
                    sku_data['Multipack Set'] = "NA"
                    
                # Align season, closure, type to closest masterdata dropdown options
                if 'season' in sku_data and sku_data['season']:
                    sku_data['season'] = get_closest_match(sku_data['season'], masterdata_choices['season'])
                    
                if 'Closure' in sku_data and sku_data['Closure']:
                    sku_data['Closure'] = get_closest_match(sku_data['Closure'], masterdata_choices['closure'])
                if 'Fastening' in sku_data and sku_data['Fastening']:
                    sku_data['Fastening'] = get_closest_match(sku_data['Fastening'], masterdata_choices['closure'])
                
                # Dresses Closure override
                if "dress" in selected_cat.lower().strip():
                    sku_data['Closure'] = "NA"
                    
                if 'Type' in sku_data and sku_data['Type']:
                    sku_data['Type'] = get_closest_match(sku_data['Type'], masterdata_choices['type'])
                
                # Override visual attributes with AI Vision predictions if available
                if vision_preds:
                    for v_col, v_val in vision_preds.items():
                        if v_val:
                            sku_data[v_col] = v_val
                        
                # Enforce "NA" for specific columns requested by the user
                for na_col in ['Sport', 'Technology', 'Add-Ons']:
                    if na_col in col_mapping:
                        sku_data[na_col] = "NA"
                        
                rows_to_insert.append(sku_data)
                inserted_count += 1
        
        # Assign sequential styleGroupId based on (vendorArticleNumber, MRP) combination changes
        current_group_id = 0
        last_seen_key = None
        for sku in rows_to_insert:
            van = sku.get('vendorArticleNumber', '')
            mrp_val = sku.get('MRP', 0)
            combine_key = (van, mrp_val)
            
            if last_seen_key is None or combine_key != last_seen_key:
                current_group_id += 1
                last_seen_key = combine_key
                
            sku['styleGroupId'] = current_group_id

        log_buffer.log("SUCCESS", f"Correlated & Prepared {inserted_count} SKU listing rows.")
        
        current_task_status["progress"] = 0.85
        current_task_status["current_step"] = "Injecting Data & Saving"
        
        # 8. Inject rows into the openpyxl template sheet
        if not dry_run:
            log_buffer.log("INFO", f"Injecting data into sheet '{target_sheet_name}' (starting at Row 4)...")
            
            current_row = 4
            for index, sku_row in enumerate(rows_to_insert):
                for key, val in sku_row.items():
                    if key in col_mapping:
                        col_idx = col_mapping[key]
                        cell = sheet.cell(row=current_row, column=col_idx)
                        cell.value = val
                        cell.font = Font(name="Segoe UI", size=10)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                current_row += 1
                
            timestamp_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
            output_filename = f"Myntra-Populated-{target_sheet_name}-{timestamp_str}.xlsx"
            output_filepath = os.path.join(output_dir, output_filename)
            
            log_buffer.log("INFO", f"Saving populated workbook to: {output_filepath}")
            wb.save(output_filepath)
            wb.close()

            # Copy to web app workspace for direct browser downloading
            workspace_dir = os.path.join("static", "temp_workspace", run_id)
            os.makedirs(workspace_dir, exist_ok=True)
            workspace_filepath = os.path.join(workspace_dir, output_filename)
            import shutil
            try:
                if os.path.abspath(output_filepath) != os.path.abspath(workspace_filepath):
                    shutil.copy(output_filepath, workspace_filepath)
            except Exception as copy_err:
                log_buffer.log("WARNING", f"Failed to save copy to web workspace: {str(copy_err)}")

            # Store file in MongoDB so it survives Railway container restarts
            try:
                import base64
                mongo_uri = None
                try:
                    from vision_classifier import get_safe_mongodb_uri
                    mongo_uri = get_safe_mongodb_uri(os.environ.get("MONGO_URI"))
                except Exception:
                    pass
                if mongo_uri:
                    import pymongo
                    _mc = pymongo.MongoClient(mongo_uri, serverSelectionTimeoutMS=8000)
                    _db = _mc.get_database()
                    with open(workspace_filepath, "rb") as _f:
                        _file_b64 = base64.b64encode(_f.read()).decode("utf-8")
                    _db["output_files"].replace_one(
                        {"run_id": run_id},
                        {"run_id": run_id, "filename": output_filename, "file_b64": _file_b64, "created_at": datetime.now().isoformat()},
                        upsert=True
                    )
                    _mc.close()
                    log_buffer.log("INFO", f"Output file persisted to cloud storage for reliable download.")
            except Exception as mongo_err:
                log_buffer.log("WARNING", f"Could not persist file to cloud storage: {str(mongo_err)}")

            log_buffer.log("SUCCESS", "=== MYNTRA LISTING GENERATION COMPLETED ===")
            log_buffer.log("SUCCESS", f"Successfully generated {inserted_count} SKUs!")
            if IS_CLOUD:
                log_buffer.log("SUCCESS", "File saved to cloud workspace. Click the 'Download Excel File' button below to download it.")
            else:
                log_buffer.log("SUCCESS", f"File saved locally to: {os.path.abspath(output_filepath)}")
            # Emit special download-trigger event picked up by the frontend in real time
            log_buffer.log("__DOWNLOAD__", json.dumps({"run_id": run_id, "filename": output_filename}))
        else:
            wb.close()
            log_buffer.log("SUCCESS", "=== DRY RUN ANALYSIS COMPLETE ===")
            log_buffer.log("SUCCESS", f"Dry run verified {inserted_count} rows matching constraints.")

        current_task_status["progress"] = 1.0
        current_task_status["status"] = "success"
        if not dry_run:
            current_task_status["output_file"] = output_filename
            current_task_status["run_id"] = run_id
        
    except Exception as e:
        log_buffer.log("ERROR", f"An error occurred: {str(e)}")
        log_buffer.log("ERROR", traceback.format_exc())
        current_task_status["status"] = "failed"
        current_task_status["error"] = str(e)

# REST API Endpoints
# Create temporary upload folder
TEMP_UPLOAD_DIR = os.path.join("static", "temp_uploads")
os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)

@app.route('/api/config')
def get_config():
    return jsonify({
        "is_cloud": IS_CLOUD,
        "has_gemini_key": bool(vision_classifier.get_api_key())
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    if file:
        filename = secure_filename(file.filename)
        unique_prefix = str(int(time.time())) + "_"
        filepath = os.path.join(TEMP_UPLOAD_DIR, unique_prefix + filename)
        file.save(filepath)
        
        # Check if zip extraction is requested
        is_zip = filename.lower().endswith('.zip')
        if is_zip and request.form.get('extract') == 'true':
            import zipfile
            extracted_dir = os.path.join(TEMP_UPLOAD_DIR, unique_prefix + "extracted")
            os.makedirs(extracted_dir, exist_ok=True)
            try:
                with zipfile.ZipFile(filepath, 'r') as zip_ref:
                    zip_ref.extractall(extracted_dir)
                return jsonify({
                    "filepath": filepath.replace("\\", "/"),
                    "extracted_dir": extracted_dir.replace("\\", "/")
                })
            except Exception as e:
                return jsonify({"error": f"Failed to extract zip file: {str(e)}"}), 500
                
        return jsonify({"filepath": filepath.replace("\\", "/")})

@app.route('/api/vision/save-key', methods=['POST'])
def vision_save_key():
    data = request.json or {}
    key = data.get("key", "")
    success = vision_classifier.save_api_key(key)
    if success:
        return jsonify({"message": "API key saved successfully!"})
    return jsonify({"error": "Failed to save API key."}), 500

@app.route('/api/vision/get-key')
def vision_get_key():
    key = vision_classifier.get_api_key()
    if key:
        masked = key[:4] + "..." + key[-4:] if len(key) > 8 else "****"
        return jsonify({"has_key": True, "masked_key": masked})
    return jsonify({"has_key": False, "masked_key": ""})

@app.route('/api/vision/dropdowns')
def vision_dropdowns():
    category = request.args.get("category", "")
    template_path = request.args.get("template_path", "")
    
    if not category:
        return jsonify({"error": "Category is required!"}), 400
        
    if not template_path or not os.path.exists(template_path):
        last_template_fallback = os.path.join(TEMP_UPLOAD_DIR, "last_template.xlsx")
        if os.path.exists(last_template_fallback):
            template_path = last_template_fallback
        else:
            template_path = "G:\\My Drive\\Manan\\Automation\\Myntra\\Myntra-Sku-Template-2026-05-20.xlsx"
            if not os.path.exists(template_path):
                template_path = "C:\\Users\\manan\\Downloads\\Myntra Learning\\Myntra-Sku-Template-2026-05-05-Top.xlsx"
                if not os.path.exists(template_path):
                    template_path = "C:\\Users\\manan\\Downloads\\Myntra-Sku-Template-2026-05-23 (1).xlsx"
                    if not os.path.exists(template_path):
                        template_path = "C:\\Users\\manan\\Downloads\\Myntra-Sku-Template-2026-05-23.xlsx"
                        if not os.path.exists(template_path):
                            return jsonify({"error": "Template file not found! Please check template settings."}), 400
                            
    try:
        fields, dropdowns = vision_classifier.get_myntra_dropdowns_and_mandatory_fields(template_path, category)
        return jsonify({
            "fields": fields,
            "dropdowns": dropdowns
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route('/api/vision/analyze', methods=['POST'])
def vision_analyze():
    category = request.form.get("category", "")
    template_path = request.form.get("template_path", "")
    
    # Check if a template file was uploaded in this request
    uploaded_template = request.files.get("template_file")
    if uploaded_template and uploaded_template.filename:
        template_path = os.path.join(TEMP_UPLOAD_DIR, "last_template.xlsx")
        uploaded_template.save(template_path)
        
    if not category:
        return jsonify({"error": "Category is required!"}), 400
        
    if not template_path or not os.path.exists(template_path):
        last_template_fallback = os.path.join(TEMP_UPLOAD_DIR, "last_template.xlsx")
        if os.path.exists(last_template_fallback):
            template_path = last_template_fallback
        else:
            # Fall back to default template
            template_path = "G:\\My Drive\\Manan\\Automation\\Myntra\\Myntra-Sku-Template-2026-05-20.xlsx"
            if not os.path.exists(template_path):
                template_path = "C:\\Users\\manan\\Downloads\\Myntra Learning\\Myntra-Sku-Template-2026-05-05-Top.xlsx"
                if not os.path.exists(template_path):
                    return jsonify({"error": "Template file not found! Please check template settings."}), 400
            
    front_file = request.files.get("front_image")
    back_file = request.files.get("back_image")
    
    if not front_file or not back_file:
        return jsonify({"error": "Both front and back images are required!"}), 400
        
    front_name = secure_filename(front_file.filename)
    back_name = secure_filename(back_file.filename)
    
    front_path = os.path.join(TEMP_UPLOAD_DIR, "temp_front_" + front_name)
    back_path = os.path.join(TEMP_UPLOAD_DIR, "temp_back_" + back_name)
    
    front_file.save(front_path)
    back_file.save(back_path)
    
    try:
        result = vision_classifier.analyze_product_images(front_path, back_path, category, template_path)
    except Exception as e:
        if os.path.exists(front_path):
            try: os.remove(front_path)
            except: pass
        if os.path.exists(back_path):
            try: os.remove(back_path)
            except: pass
        return jsonify({"error": str(e)}), 400
        
    if "error" in result:
        # Clean up
        if os.path.exists(front_path):
            try: os.remove(front_path)
            except: pass
        if os.path.exists(back_path):
            try: os.remove(back_path)
            except: pass
        return jsonify(result), 400
        
    return jsonify({
        "predictions": result["predictions"],
        "dropdowns": result["dropdowns"],
        "front_temp_path": front_path.replace("\\", "/"),
        "back_temp_path": back_path.replace("\\", "/")
    })

@app.route('/api/vision/save-learning', methods=['POST'])
def vision_save_learning():
    data = request.json or {}
    van = data.get("vendorArticleNumber", "").strip()
    category = data.get("category", "").strip()
    front_temp = data.get("front_temp_path", "")
    back_temp = data.get("back_temp_path", "")
    labels = data.get("labels", {})
    
    if not van or not category:
        return jsonify({"error": "SKU code and category are required!"}), 400
        
    success = vision_classifier.save_learning(van, category, front_temp, back_temp, labels)
    
    # Clean up temp files
    if front_temp and os.path.exists(front_temp) and "temp_uploads" in front_temp:
        try: os.remove(front_temp)
        except: pass
    if back_temp and os.path.exists(back_temp) and "temp_uploads" in back_temp:
        try: os.remove(back_temp)
        except: pass
        
    if success:
        return jsonify({"message": "Learning saved & Model updated successfully!"})
    return jsonify({"error": "Failed to save learning."}), 500

@app.route('/api/vision/get-learnings')
def vision_get_learnings():
    learnings = vision_classifier.get_learnings()
    mapped_learnings = []
    for k, v in learnings.items():
        doc = v.copy()
        doc["front_image"] = v.get("front_temp_path", "")
        doc["back_image"] = v.get("back_temp_path", "")
        mapped_learnings.append(doc)
    return jsonify({"learnings": mapped_learnings})

@app.route('/api/vision/delete-learning/<vendor_article_num>', methods=['DELETE'])
def vision_delete_learning(vendor_article_num):
    success = vision_classifier.delete_learning(vendor_article_num)
    if success:
        return jsonify({"message": "Learning deleted successfully."})
    return jsonify({"error": "Failed to delete learning."}), 500

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/categories')
def get_categories():
    return jsonify({
        "categories": MYNTRA_ARTICLE_TYPES,
        "defaults": CATEGORY_DEFAULTS
    })

@app.route('/api/browse', methods=['POST'])
def browse_path():
    if IS_CLOUD or tk is None or filedialog is None:
        return jsonify({"error": "Local file browsing is disabled in Cloud Mode"}), 400
    data = request.json or {}
    browse_type = data.get('type', 'file') # 'file' or 'folder'
    title = data.get('title', 'Select File')
    initial_dir = data.get('initialdir', '')
    
    # Run Tkinter browser inside a thread to prevent blocking the Flask server
    result_queue = queue.Queue()
    
    def tk_selector():
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        if browse_type == 'folder':
            path = filedialog.askdirectory(title=title, initialdir=initial_dir)
        else:
            path = filedialog.askopenfilename(title=title, initialdir=initial_dir, filetypes=[("Excel Files", "*.xlsx *.xls")])
        root.destroy()
        result_queue.put(path)
        
    t = threading.Thread(target=tk_selector)
    t.start()
    t.join()
    
    selected_path = result_queue.get()
    return jsonify({"path": selected_path})

@app.route('/api/status')
def get_status():
    return jsonify(current_task_status)

@app.route('/api/run', methods=['POST'])
def run_automation():
    global current_task_status
    if current_task_status["status"] == "running":
        return jsonify({"error": "A task is already running!"}), 400
        
    data = request.json or {}
    item_path = data.get('item_path')
    content_path = data.get('content_path')
    template_path = data.get('template_path')
    output_dir = data.get('output_dir')
    product_images_dir = data.get('product_images_dir', '')
    
    category = data.get('category', 'Shorts')
    isp_rule = data.get('isp_rule', 'Discounted')
    isp_discount = float(data.get('isp_discount', 50.0))
    products_filter = data.get('products_filter', '')
    use_vision = data.get('use_vision', False)
    dry_run = data.get('dry_run', False)
    
    run_id = str(int(time.time()))
    if IS_CLOUD:
        output_dir = os.path.join("static", "temp_workspace", run_id)
        os.makedirs(output_dir, exist_ok=True)
            
    if not item_path or not template_path or (not IS_CLOUD and not output_dir):
        return jsonify({"error": "Item Directory and SKU Template are required!" if IS_CLOUD else "Item Directory, SKU Template, and Output Save Folder are required!"}), 400
        
    # Save a cached copy of the template for the AI Vision tab dropdown list query
    import shutil
    try:
        os.makedirs("static/temp_uploads", exist_ok=True)
        shutil.copy(template_path, os.path.join("static", "temp_uploads", "last_template.xlsx"))
    except Exception as copy_err:
        print(f"Failed to copy template to last_template.xlsx: {str(copy_err)}")
    params = {
        "run_id": run_id,
        "item_path": item_path,
        "content_path": content_path,
        "template_path": template_path,
        "output_dir": output_dir,
        "category": category,
        "isp_rule": isp_rule,
        "isp_discount": isp_discount,
        "products_filter": products_filter,
        "use_vision": use_vision,
        "product_images_dir": product_images_dir
    }
    
    # Reset status
    current_task_status = {
        "status": "running",
        "progress": 0.0,
        "current_step": "Initializing background worker",
        "error": None,
        "run_id": run_id
    }
    
    # Launch processing in background thread
    t = threading.Thread(target=run_automation_core, args=(params, dry_run), daemon=True)
    t.start()
    
    return jsonify({"message": "Job initiated successfully"})

@app.route('/api/vision/download-output')
def download_output():
    run_id = request.args.get("run_id", "")
    filename = request.args.get("filename", "")
    if not run_id or not filename:
        return "Missing parameters", 400
        
    directory = os.path.join("static", "temp_workspace", run_id)
    if not os.path.exists(os.path.join(directory, filename)):
        return f"File '{filename}' not found in run '{run_id}'", 404
        
    return send_from_directory(directory, filename, as_attachment=True)

@app.route('/api/vision/export-db')
def export_db():
    if not os.path.exists("myntra_vision_learning.json"):
        return jsonify({"error": "No trained database found."}), 404
    return send_from_directory(".", "myntra_vision_learning.json", as_attachment=True)

@app.route('/api/logs')
def stream_logs():
    q = log_buffer.add_listener()

    def event_stream():
        try:
            yield f"data: {json.dumps({'time': datetime.now().strftime('%H:%M:%S'), 'level': 'INFO', 'message': 'Web Terminal Connected'})}\n\n"
            while True:
                try:
                    log_entry = q.get(timeout=20.0)
                    yield f"data: {json.dumps(log_entry)}\n\n"
                except queue.Empty:
                    # Send a real SSE data event (not a comment) so Railway's proxy
                    # sees active traffic and does not close the connection
                    yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
        except GeneratorExit:
            pass
        finally:
            log_buffer.remove_listener(q)

    response = Response(event_stream(), mimetype="text/event-stream")
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['Connection'] = 'keep-alive'
    response.headers['X-Accel-Buffering'] = 'no'
    return response


@app.route('/api/download-file')
def download_file_from_db():
    """Download output file — reads from MongoDB if not on local disk (handles Railway restarts)."""
    run_id = request.args.get("run_id", "")
    filename = request.args.get("filename", "")
    if not run_id or not filename:
        return "Missing parameters", 400

    # Try local filesystem first
    directory = os.path.join("static", "temp_workspace", run_id)
    local_path = os.path.join(directory, filename)
    if os.path.exists(local_path):
        return send_from_directory(directory, filename, as_attachment=True)

    # Fallback: fetch from MongoDB
    try:
        from vision_classifier import get_safe_mongodb_uri
        mongo_uri = get_safe_mongodb_uri(os.environ.get("MONGO_URI"))
        if mongo_uri:
            import pymongo, base64
            client = pymongo.MongoClient(mongo_uri, serverSelectionTimeoutMS=8000)
            db = client.get_database()
            doc = db["output_files"].find_one({"run_id": run_id, "filename": filename})
            client.close()
            if doc and doc.get("file_b64"):
                file_bytes = base64.b64decode(doc["file_b64"])
                os.makedirs(directory, exist_ok=True)
                with open(local_path, "wb") as f:
                    f.write(file_bytes)
                return send_from_directory(directory, filename, as_attachment=True)
    except Exception as e:
        print(f"Failed to fetch file from MongoDB: {e}")

    return f"File '{filename}' not found. The server may have restarted.", 404

if __name__ == '__main__':
    # Make sure static directory exists
    os.makedirs('static', exist_ok=True)
    print("====================================================")
    print("  MYNTRA AUTO LISTER PRO - WEB SERVER STARTED       ")
    print("  Open your browser and navigate to:                 ")
    print("  http://127.0.0.1:5000                             ")
    print("====================================================")
    port = int(os.environ.get("PORT", 5000))
    host = os.environ.get("HOST", "127.0.0.1")
    app.run(host=host, port=port, debug=False)
