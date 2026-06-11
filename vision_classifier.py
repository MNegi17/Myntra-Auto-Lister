import os
import json
import openpyxl
from PIL import Image
import base64

KEY_FILE = "gemini_api_key.txt"
LEARNINGS_FILE = "myntra_vision_learning.json"

def file_to_base64_data_url(filepath):
    if not filepath:
        return ""
    if str(filepath).startswith("data:image/"):
        return filepath
    if not os.path.exists(filepath):
        return filepath
    try:
        ext = os.path.splitext(filepath)[1].lower().replace(".", "")
        if not ext:
            ext = "jpeg"
        elif ext == "jpg":
            ext = "jpeg"
        with open(filepath, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode("utf-8")
            return f"data:image/{ext};base64,{encoded_string}"
    except Exception as e:
        print(f"Failed to encode image to base64: {str(e)}")
    return filepath

def save_api_key(key):
    try:
        with open(KEY_FILE, "w", encoding="utf-8") as f:
            f.write(key.strip())
        return True
    except:
        return False

def get_api_key():
    env_key = os.environ.get("GEMINI_API_KEY")
    if env_key:
        return env_key
    if os.path.exists(KEY_FILE):
        try:
            with open(KEY_FILE, "r", encoding="utf-8") as f:
                return f.read().strip()
        except:
            pass
    return ""

def get_learnings():
    mongo_uri = os.environ.get("MONGO_URI")
    if mongo_uri:
        try:
            import pymongo
            client = pymongo.MongoClient(mongo_uri)
            db = client.get_database()
            collection = db.learnings
            learnings = {}
            for doc in collection.find():
                van = doc.get("vendorArticleNumber")
                if van:
                    learnings[str(van).upper()] = {
                        "vendorArticleNumber": van,
                        "category": doc.get("category"),
                        "front_temp_path": doc.get("front_temp_path"),
                        "back_temp_path": doc.get("back_temp_path"),
                        "labels": doc.get("labels", {})
                    }
            client.close()
            return learnings
        except Exception as e:
            print(f"Failed to read from MongoDB: {str(e)}")

    if os.path.exists(LEARNINGS_FILE):
        try:
            with open(LEARNINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {}

def save_learning(*args, **kwargs):
    # Robust argument parsing to support all calling signatures
    van = kwargs.get("vendor_article_num") or kwargs.get("van")
    category = kwargs.get("category")
    front = kwargs.get("front_img_path") or kwargs.get("front_temp") or kwargs.get("front_temp_path")
    back = kwargs.get("back_img_path") or kwargs.get("back_temp") or kwargs.get("back_temp_path")
    labels = kwargs.get("labels")
    
    if len(args) >= 5:
        van = van or args[0]
        category = category or args[1]
        front = front or args[2]
        back = back or args[3]
        labels = labels or args[4]
    elif len(args) == 4:
        van = van or args[0]
        category = category or args[1]
        front = front or args[2]
        back = back or args[3]

    if not van:
        return False
        
    front_b64 = file_to_base64_data_url(front)
    back_b64 = file_to_base64_data_url(back)
    
    mongo_uri = os.environ.get("MONGO_URI")
    if mongo_uri:
        try:
            import pymongo
            client = pymongo.MongoClient(mongo_uri)
            db = client.get_database()
            collection = db.learnings
            doc = {
                "vendorArticleNumber": van,
                "category": category,
                "front_temp_path": front_b64,
                "back_temp_path": back_b64,
                "labels": labels or {}
            }
            collection.replace_one(
                {"vendorArticleNumber": {"$regex": f"^{van}$", "$options": "i"}},
                doc,
                upsert=True
            )
            client.close()
            return True
        except Exception as e:
            print(f"Failed to save to MongoDB: {str(e)}")
            return False

    learnings = get_learnings()
    learnings[str(van).upper()] = {
        "vendorArticleNumber": van,
        "category": category,
        "front_temp_path": front_b64,
        "back_temp_path": back_b64,
        "labels": labels or {}
    }
    
    try:
        with open(LEARNINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(learnings, f, indent=4)
        return True
    except:
        return False

def delete_learning(vendor_article_num):
    mongo_uri = os.environ.get("MONGO_URI")
    if mongo_uri:
        try:
            import pymongo
            client = pymongo.MongoClient(mongo_uri)
            db = client.get_database()
            collection = db.learnings
            res = collection.delete_one({"vendorArticleNumber": {"$regex": f"^{vendor_article_num}$", "$options": "i"}})
            client.close()
            return res.deleted_count > 0
        except Exception as e:
            print(f"Failed to delete from MongoDB: {str(e)}")
            return False

    learnings = get_learnings()
    key = str(vendor_article_num).upper()
    if key in learnings:
        del learnings[key]
        try:
            with open(LEARNINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(learnings, f, indent=4)
            return True
        except:
            pass
    return False

def get_myntra_dropdowns_and_mandatory_fields(template_path, category):
    if not os.path.exists(template_path):
        return [], {}
        
    try:
        wb = openpyxl.load_workbook(template_path, read_only=True)
    except Exception as e:
        return [], {}
        
    # Find matching category sheet
    target_sheet = None
    category_sheets = [s for s in wb.sheetnames if s not in ["__INSTRUCTIONS", "masterdata"]]
    for sheet_name in wb.sheetnames:
        if sheet_name not in ["__INSTRUCTIONS", "masterdata"]:
            if sheet_name.lower().strip() == category.lower().strip():
                target_sheet = sheet_name
                break
    if not target_sheet and category_sheets:
        target_sheet = category_sheets[0]
        
    if not target_sheet:
        wb.close()
        return [], {}
        
    sheet = wb[target_sheet]
    rows = list(sheet.iter_rows(max_row=3, values_only=True))
    wb.close()
    
    if len(rows) < 3:
        return [], {}
        
    display_names = rows[1]
    tech_keys = rows[2]
    
    fields = []
    dropdowns = {}
    
    # Compile the fields
    for d_name, t_key in zip(display_names, tech_keys):
        if t_key and d_name:
            field_name = str(t_key).strip()
            is_mandatory = "*" in str(d_name)
            fields.append({
                "name": field_name,
                "label": str(d_name).replace("*", "").strip(),
                "mandatory": is_mandatory
            })
            
    # Compile the masterdata validation dropdowns
    try:
        wb = openpyxl.load_workbook(template_path, data_only=True)
        if "masterdata" in wb.sheetnames:
            m_sheet = wb["masterdata"]
            for col in m_sheet.iter_cols(values_only=True):
                if col and col[0]:
                    header_str = str(col[0]).strip()
                    vals = [str(val).strip() for val in col[1:] if val is not None]
                    if vals:
                        dropdowns[header_str] = sorted(list(set(vals)))
        wb.close()
    except:
        pass
        
    return fields, dropdowns

def analyze_product_images(front_path, back_path, category, template_path):
    api_key = get_api_key()
    if not api_key:
        return {"error": "Gemini API key is not configured. Please save your API key in Settings."}
        
    fields, dropdowns = get_myntra_dropdowns_and_mandatory_fields(template_path, category)
    if not fields:
        return {"error": f"Could not load columns for category '{category}' from the sheet template."}
        
    try:
        import google.generativeai as genai
    except ImportError:
        return {"error": "Google Generative AI package is not installed. Run 'pip install google-generativeai'."}
        
    # Build a structured prompt describing technical schema and allowed choices
    prompt = f"""
You are an expert fashion cataloging AI assistant. Analyze the attached front and back images of a children's garment of category '{category}'.
Based on the visual features, classify the garment according to the following schema of technical attributes.
For attributes with dropdown options, you MUST choose exactly one option from the allowed values provided below.

Attributes to predict:
"""
    for f in fields:
        name = f["name"]
        label = f["label"]
        mandatory = "Mandatory" if f["mandatory"] else "Optional"
        
        # Find closest match in masterdata dropdown keys
        options = []
        for key in dropdowns:
            if key.lower() in name.lower() or name.lower() in key.lower():
                options = dropdowns[key]
                break
                
        prompt += f"- {name} ({label}): {mandatory}. "
        if options:
            prompt += f"Allowed values: {options}\n"
        else:
            prompt += "Free text/description based.\n"
            
    prompt += """
Return the predictions as a valid JSON object mapping technical attribute names to their predicted values.
Do not wrap the JSON in ```json or any other text; output raw JSON only.
"""

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.5-flash')
        
        front_img = Image.open(front_path)
        back_img = Image.open(back_path)
        
        response = model.generate_content([prompt, front_img, back_img])
        
        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
        
        predictions = json.loads(text)
        
        clean_preds = {}
        for f in fields:
            name = f["name"]
            clean_preds[name] = predictions.get(name, "")
            
        return {
            "predictions": clean_preds,
            "dropdowns": dropdowns
        }
    except Exception as e:
        return {"error": f"Gemini API invocation failed: {str(e)}"}
